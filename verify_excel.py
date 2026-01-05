from openpyxl import load_workbook

# 读取Excel文件
wb = load_workbook('book_list-python.xlsx')
ws = wb.active

print("Excel文件前10行内容：")
print("=" * 100)

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 10:  # 只显示前10行
        break
    print(f"{i}. {row}")

print("=" * 100)
print(f"\n总共 {ws.max_row - 1} 行数据（不含表头）")
