# -*- coding: UTF-8 -*-
"""
豆瓣书籍爬虫 - Python 3版本
从原始Python 2代码迁移而来
"""

import sys
import time
import urllib.parse
import urllib.request
import random
import numpy as np
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Some User Agents
hds = [
    {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
    {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
    {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}
]


def book_spider(book_tag):
    """爬取指定标签的豆瓣书籍信息"""
    page_num = 0
    book_list = []
    try_times = 0

    while True:
        # 构造URL
        url = 'https://book.douban.com/tag/' + urllib.parse.quote(book_tag) + '?start=' + str(page_num * 20)

        # 随机延迟，避免被封
        time.sleep(random.random() * 5)

        try:
            req = urllib.request.Request(url, headers=hds[page_num % len(hds)])
            source_code = urllib.request.urlopen(req, timeout=10).read()
            plain_text = str(source_code, 'utf-8')
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(f"HTTP/URL错误: {e}")
            break
        except Exception as e:
            print(f"请求错误: {e}")
            break

        soup = BeautifulSoup(plain_text, 'html.parser')
        list_soup = soup.find('div', {'class': 'mod book-list'})

        try_times += 1
        if list_soup is None and try_times < 10:
            continue
        elif list_soup is None or len(list_soup) <= 1:
            break  # Break when no information got

        for book_info in list_soup.findAll('dd'):
            try:
                title = book_info.find('a', {'class': 'title'}).string.strip()
            except:
                title = '暂无书名'

            try:
                desc = book_info.find('div', {'class': 'desc'}).string.strip()
                desc_list = desc.split('/')
            except:
                desc_list = ['暂无']

            try:
                book_url = book_info.find('a', {'class': 'title'}).get('href')
            except:
                book_url = ''

            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者： 暂无'

            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'

            try:
                rating = book_info.find('span', {'class': 'rating_nums'}).string.strip()
            except:
                rating = '0.0'

            try:
                people_num = book_info.findAll('span')[2].string.strip()
                people_num = people_num.strip('人评价')
            except:
                people_num = '0'

            book_list.append([title, rating, people_num, author_info, pub_info])
            try_times = 0  # set 0 when got valid information

        page_num += 1
        print(f'正在下载第 {page_num} 页信息')

    return book_list


def get_people_num(url):
    """获取书籍评价人数"""
    try:
        req = urllib.request.Request(url, headers=hds[random.randint(0, len(hds) - 1)])
        source_code = urllib.request.urlopen(req, timeout=10).read()
        plain_text = str(source_code, 'utf-8')
    except Exception as e:
        print(f"获取评价人数时出错: {e}")
        return '0'

    soup = BeautifulSoup(plain_text, 'html.parser')
    try:
        people_num = soup.find('div', {'class': 'rating_sum'}).findAll('span')[1].string.strip()
        return people_num
    except:
        return '0'


def do_spider(book_tag_lists):
    """执行爬虫主程序"""
    book_lists = []
    for book_tag in book_tag_lists:
        print(f"\n开始爬取标签: {book_tag}")
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: float(x[1]), reverse=True)
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists, book_tag_lists):
    """将爬取结果保存到Excel文件"""
    wb = Workbook(write_only=True)
    ws = []

    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i]))

    for i in range(len(book_tag_lists)):
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
        count = 1
        for bl in book_lists[i]:
            try:
                ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            except:
                ws[i].append([count, bl[0], 0.0, 0, bl[3], bl[4]])
            count += 1

    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i])
    save_path += '.xlsx'

    wb.save(save_path)
    print(f"\n结果已保存到: {save_path}")


if __name__ == '__main__':
    # 测试标签列表（从小规模开始）
    book_tag_lists = ['python']

    print("=" * 50)
    print("豆瓣书籍爬虫 - Python 3版本")
    print("=" * 50)

    book_lists = do_spider(book_tag_lists)

    if book_lists and book_lists[0]:
        print_book_lists_excel(book_lists, book_tag_lists)
        print(f"\n爬取完成！共获取 {len(book_lists[0])} 本书籍信息")
    else:
        print("\n爬取失败或未获取到数据，可能原因：")
        print("1. 豆瓣网站结构已改变，需要重新分析页面")
        print("2. IP被暂时限制访问")
        print("3. 网络连接问题")
